from logs.logs import p_log
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def write_2dlist_to_excel(diff_list, save_dir):
    def escape_excel_formulas(value):
        """Экранирует строки, начинающиеся с '=', добавляя апостроф."""
        if isinstance(value, str) and value.startswith('='):
            return "'" + value
        return value

    def set_column_widths(sheet):
        """Устанавливает ширину колонок по максимальной длине содержимого."""
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)

            # Определяем максимальную длину содержимого в колонке
            for cell in col:
                try:
                    # Для чисел учитываем форматирование с пробелами
                    if isinstance(cell.value, (int, float)):
                        formatted = "{:,.0f}".format(cell.value).replace(",", " ")
                        cell_length = len(formatted)
                    else:
                        cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length
                except Exception as err:
                    p_log(f"Ошбика определения максимальной длины ячейки: {err}")
                    continue

            # Устанавливаем ширину с небольшим запасом
            sheet.column_dimensions[column].width = max_length + 7

    def add_data_to_sheet(sheet, data, header):
        """Добавляет данные и заголовки в лист Excel."""
        # Добавляем заголовки
        sheet.append(header)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')  # Выравнивание заголовков по центру

        # Добавляем данные
        for row in data:
            safe_row = [escape_excel_formulas(item) for item in row]
            sheet.append(safe_row)

        # Заливка розовым для ордена [SLAVS] (3-я колонка)
        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[2].value == "[SLAVS]":  # Индекс 2 = колонка "Орден"
                for cell in row:
                    cell.fill = pink_fill

        # Устанавливаем выравнивание для данных
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for col_idx, cell in enumerate(row, start=1):
                if col_idx in (5, 7) and isinstance(cell.value, (int, float)):
                    # Формат: 6 448 011 (для любых чисел)
                    cell.number_format = '# ### ### ##0' if cell.value >= 1e6 else '# ##0'
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')

        # set_column_widths(sheet)

    # Сортируем данные перед записью (по "Добыча UP" — 7-й элемент в строке)
    try:
        # Фильтруем только строки с числами в колонке "Добыча UP"
        sortable_data = [row for row in diff_list if isinstance(row[6], (int, float))]
        # Сортируем по убыванию
        sorted_data = sorted(sortable_data, key=lambda x: -x[6])
        # Добавляем оставшиеся строки (где нет чисел)
        other_data = [row for row in diff_list if not isinstance(row[6], (int, float))]
        final_data = sorted_data + other_data
    except (IndexError, TypeError) as er:
        p_log(f"Ошибка сортировки: {er}")
        final_data = diff_list  # Если что-то пошло не так, оставляем как есть

    # Создаем объект Workbook и листы
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Статистика с 24.07.24"
    header1 = ['ID', 'Имя', 'Орден', 'Уровень', 'Добыча', 'Уровень UP', 'Добыча UP', 'Используемые имена',
               'Был в ордене']

    add_data_to_sheet(sheet1, final_data, header1)

    # Устанавливаем ширину колонок после добавления всех данных
    set_column_widths(sheet1)

    # Сохраняем рабочую книгу
    workbook.save(save_dir)
    workbook.close()

    p_log(f"Список успешно сохранен в файл {save_dir}")
