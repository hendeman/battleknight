import re
import pickle
import os
import pandas as pd
from time import sleep
from datetime import datetime

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from bs4 import BeautifulSoup

from logs.logs import p_log
from module.all_function import all_party
from module.data_pars import visit, party
from setting import exclusion_list, url_members, clan_html_file, FILE_NAME, deco_func, url_gold, SERVER

from module.all_function import day, syntax_day, create_folder
from module.data_pars import pars_player
from module.http_requests import make_request, post_request
from setting import url_stat, STAT_FILE_NAME, folder_name, STAT_FILE_LOSS, folder_name_loss, today

GOLD_DAY = 100
DATA_CHANGE_FILE = datetime.now()


# ____________________________________________ Статистика ордена ________________________________________________

def exclude_keys_decorator(exclusion_list=None, deco_func=False):
    def decorator(func):
        def wrapper(dict1, dict2):
            dict3 = func(dict1, dict2)
            if exclusion_list:
                for key in exclusion_list:
                    dict3.pop(key, None)
            return dict3

        return wrapper if deco_func else func  # Возвращаем обертку или саму функцию

    return decorator


# Декоратор exclude_keys_decorator добавляет список исключения exclusion_list = ["Ksusha"]
@exclude_keys_decorator(exclusion_list=exclusion_list, deco_func=deco_func)
def replenish_treasury(a: dict, b: dict) -> dict:
    # Функция для обновления ордена
    def update_order(a, b, add_message=None, remove_message=None):
        if add_message:
            knight_names = [a[x]["name"] for x in set(a).difference(set(b))]
            p_log(f'Добро пожаловать в орден {", ".join(knight_names)}')
        if remove_message:
            knight_names = [b[x]["name"] for x in set(b).difference(set(a))]
            p_log(f'Рыцарь {", ".join(knight_names)} вышел из ордена')

        for i in set(a).difference(set(b)):
            b[i] = a[i]
            b[i]["gold"] = 0

    # Проверка и обновление ордена
    if len(a) < len(b):
        update_order(a, b, None, 'Рыцарь вышел из ордена')

    elif len(a) > len(b):
        update_order(a, b, 'Добро пожаловать в орден', None)

    elif len(set(a).intersection(set(b))) < len(b):
        update_order(a, b, 'Добро пожаловать в орден', 'Рыцарь вышел из ордена')

    return {value["name"]: {"gold": (lambda x, y: x - y if x - y >= 0 else x)(value["gold"], b[key]["gold"]),
                            "level": value["level"]}
            for key, value in a.items()
            if value["time"] <= 4}


def get_gold_day():
    global GOLD_DAY
    resp = make_request(url_gold)
    pattern = r'"payments":(\d+)'
    match = re.search(pattern, resp.text)
    if match:
        payments_value = int(match.group(1))  # Получаем полное соответствие
        GOLD_DAY = payments_value


def gold_factor(a: dict, pas_day: int) -> dict:
    pas_day = 1 if pas_day == 0 else pas_day
    total_levels_sum = sum(map(lambda x: x['level'], a.values()))
    factor = round(GOLD_DAY / total_levels_sum, 2)
    return {key: {'gold': value['gold'],
                  'kg': int((value["gold"] * 100 / pas_day) / (value["level"] * factor)),
                  'balance': int(value['gold'] - value["level"] * factor * pas_day),
                  'level': value["level"]} for
            key, value in a.items()}


def print_data(ss, debet, credit, days_have_passed, write_flag):
    pref = "_all" if write_flag else ""
    txt_report = f"bk\\report_clan\\report_{today.day:02d}_{today.month:02d}_{today.year}{pref}.txt"
    try:
        os.makedirs(os.path.dirname(txt_report), exist_ok=True)
        with open(txt_report, 'w', encoding='utf-8') as file:
            file.write(f"""Статистика пополнения казны за {days_have_passed} {syntax_day(days_have_passed)}.
Всего сдано {debet} золота. Потрачено куклой {credit} золота.
Итого: {'+' if debet - credit >= 0 else ''}{debet - credit}
Меценатом недели становится {list(ss.keys())[0]} !!!\n""")
            for i in ss.items():
                file.write(
                    f"{i[1]['gold']:6} | {['', '+'][i[1]['balance'] >= 0]}{i[1]['balance']} золота → {i[0]} {i[1]['level']}ур. ({i[1]['kg']} %)\n")
            file.write("""Всего пополнил в казну | +избыток, -недосдача → имя (KZ %)
KZ - соотношение между внесенным в казну золотом и необходимым (в процентах)""")
            p_log(f"Отчет {txt_report} успешно создан")
    except Exception as er:
        p_log(f"Ошибка записи {txt_report}: {er}", level='warning')


def get_statistic_clan(write_flag):
    resp = make_request(url_members)
    get_gold_day()  # Обновить переменную содержание замка в день
    os.makedirs(os.path.dirname(clan_html_file), exist_ok=True)
    with open(clan_html_file, 'w', encoding='utf-8') as file:
        file.write(resp.text)

    with open(clan_html_file, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'lxml')

        time_visit_number, party_gold = visit(soup), party(soup)

        if len(time_visit_number) != len(party_gold):
            raise f"Нет совпадения длины списков: {len(time_visit_number)}!={len(party_gold)}"
        all_dct = all_party(time_visit_number, party_gold)

        with open(FILE_NAME, 'rb') as f:
            loaded_dict = pickle.load(f)
            days_have_passed = day(FILE_NAME)  # количество дней для статистики int

            dc = gold_factor(replenish_treasury(all_dct, loaded_dict),
                             days_have_passed)  # Коэфф. Золота(сравнить два словаря)

            ss = dict(sorted(dc.items(), key=lambda item: item[1]["balance"], reverse=True))
            debet, credit = sum(map(lambda x: x['gold'], ss.values())), days_have_passed * GOLD_DAY

            print_data(ss, debet, credit, days_have_passed,
                       write_flag)  # вывести данные в консоль, создание отчета report.txt

        return all_dct


# _____________________________________ Статистика онлайна _____________________________________________
def get_statistic() -> dict:
    stat_dct = {}
    create_folder(folder_name)  # создание папки
    make_request(url_stat)

    p_log("Прогресс: ....")

    for i in range(0, 2000, 100):
        param = {
            'highscoreOffset': str(i),  # выберите нужное значение
            'sort': 'loot',
            'searchUser': ''  # укажите значение, если необходимо
        }
        stat = post_request(url_stat, param)
        with open(f'{folder_name}\\{i + 100}_BattleKnight.html', 'w', encoding='utf-8') as file:
            file.write(stat.text)
        soup = BeautifulSoup(stat.text, 'lxml')
        stat_dct.update(pars_player(soup))
        sleep(3)
    p_log(f"Ожидание паузы в 30 секунд перед парсингом потерь игроков...")
    sleep(30)
    return stat_dct


def dict_values_difference(pars_dct: dict) -> list:
    with open(STAT_FILE_NAME, 'rb') as file1:
        loaded_dict = pickle.load(file1)
        global DATA_CHANGE_FILE
        DATA_CHANGE_FILE = datetime.fromtimestamp(os.path.getmtime(STAT_FILE_NAME))
        days_have_passed = day(STAT_FILE_NAME)
        p_log(f'Статистика за {days_have_passed} {syntax_day(days_have_passed)}')

        nested_list = []
        dc = {}
        create_folder(folder_name_loss)  # создание папки

        p_log("Сканирование убытка играющих рыцарей:")
        for key1 in pars_dct.keys() & loaded_dict.keys():
            if pars_dct[key1]['gold'] - loaded_dict[key1]['gold'] > 1000 or pars_dct[key1]['victory'] - \
                    loaded_dict[key1]['victory'] > 10:
                print("*", end="")
                url = f'{SERVER}/common/profile/{key1}/Scores/Player'
                resp = make_request(url, game_sleep=False)
                with open(f'{folder_name_loss}\\{key1}_{pars_dct[key1]["name"]}.html', 'w', encoding='utf-8') as file2:
                    file2.write(resp.text)

                soup = BeautifulSoup(resp.text, 'lxml')
                a = soup.find('table', class_='profileTable').find_all('tr')[4]
                dc[key1] = {"loss": int(a.text.split()[2])}
                with open(STAT_FILE_LOSS, 'rb') as f:
                    loss_dict = pickle.load(f)
                    value_loss = 0 if key1 not in loss_dict else dc[key1]["loss"] - loss_dict[key1]["loss"]
                value1, value2 = pars_dct[key1], loaded_dict[key1]
                profit = value1['gold'] - value2['gold'] if value1['gold'] - value2['gold'] > 0 else 1
                nested_list.append([value1['name'],
                                    value1['clan'],
                                    value1['level'],
                                    value1['gold'] - value2['gold'],
                                    value_loss,
                                    round((100 * value_loss) / profit, 2),
                                    value1['fights'] - value2['fights'],
                                    value1['victory'] - value2['victory'],
                                    value1['defeats'] - value2['defeats']
                                    ])
                sleep(3)
        print()
        p_log("Сканирование завершено")
        return [nested_list, dc]


def write_2dlist_to_excel(diff_list, write_flag):
    pref = "_all" if write_flag else ""
    excel_file_path = f"bk\\result_xlsx\\stat_{today.day:02d}_{today.month:02d}_{today.year}{pref}.xlsx"

    def set_column_widths(sheet):
        """Устанавливает ширину колонок на основе длины данных."""
        column_widths = [len(str(cell_value)) for row in sheet.iter_rows() for cell_value in row]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[
                get_column_letter(i)].width = column_width + 1  # Добавляем дополнительные пиксели для промежутка

    def add_data_to_sheet(sheet, period, data, header):
        """Добавляет данные и заголовки в лист Excel."""
        sheet.append(period)
        sheet.append(header)
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        for row in data:
            sheet.append(row)
        set_column_widths(sheet)

    # Создаем объект Workbook и листы
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "knight"

    # Добавляем данные на первый лист
    days_have_passed = day(STAT_FILE_NAME)
    data_stat = [f'Статистика за {days_have_passed} {syntax_day(days_have_passed)}:',
                 f'c {DATA_CHANGE_FILE.strftime("%d.%m.%Y %H:%M")}',
                 f'по {datetime.now().strftime("%d.%m.%Y %H:%M")}']
    header1 = ['Имя', 'Орден', 'Уровень', 'Добыча', 'Потери', 'пот/доб (%)', 'Бои', 'Победы', 'Поражения']

    add_data_to_sheet(sheet1, data_stat, diff_list, header1)

    # Обработка данных с использованием pandas
    df = pd.DataFrame(diff_list, columns=header1)
    df['Орден'] = df['Орден'].replace('', 'no orden')
    grouped_df = df.groupby('Орден').agg({'Добыча': 'sum', 'Потери': 'sum'}).reset_index()

    total_dobych = grouped_df['Добыча'].sum()
    grouped_df['доб(%)'] = ((grouped_df['Добыча'] / total_dobych) * 100).round(2)
    grouped_df['пот/доб (%)'] = (grouped_df['Потери'] / grouped_df['Добыча'] * 100).round(2)

    result_list = grouped_df.values.tolist()

    # Создаем второй лист и добавляем данные
    sheet2 = workbook.create_sheet(title="castles")
    header2 = ['Орден', 'Добыча', 'Потери', 'доб(%)', 'пот/доб (%)']
    add_data_to_sheet(sheet2, data_stat, result_list, header2)

    # Сохраняем рабочую книгу
    workbook.save(excel_file_path)

    p_log(f"Список успешно сохранен в файл {excel_file_path}")
