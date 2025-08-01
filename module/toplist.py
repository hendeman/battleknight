from time import sleep
import os
from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from bs4 import BeautifulSoup

from module.all_function import day, syntax_day, create_folder
from module.data_pars import pars_player
from module.http_requests import make_request, post_request
from setting import url_stat, STAT_FILE_NAME, folder_name, STAT_FILE_LOSS, \
    folder_name_loss, today

DATA_CHANGE_FILE = datetime.now()


def get_statistic() -> dict:
    stat_dct = {}
    create_folder(folder_name)  # создание папки
    make_request(url_stat)

    print("*" * 20, " Прогресс: *обработка, всего 20")

    for i in range(0, 2000, 100):
        print("*", end='')
        param = {
            'highscoreOffset': str(i),  # выберите нужное значение
            'sort': 'loot',
            'searchUser': ''  # укажите значение, если необходимо
        }
        stat = post_request(url_stat, param)
        with open(f'{folder_name}\\{i + 100}_BattleKnight.html', 'w', encoding='utf-8') as file:
            file.write(stat.text)
        soup = BeautifulSoup(stat.text, 'lxml')
        stat_dct.update(pars_player(soup))
        sleep(3)
    print()
    print(f"Ожидание паузы в 30 секунд перед парсингом потерь игроков...")
    sleep(30)
    return stat_dct


def dict_values_difference(pars_dct: dict) -> list:
    with open(STAT_FILE_NAME, 'rb') as file1:
        loaded_dict = pickle.load(file1)
        global DATA_CHANGE_FILE
        DATA_CHANGE_FILE = datetime.fromtimestamp(os.path.getmtime(STAT_FILE_NAME))
        days_have_passed = day(STAT_FILE_NAME)
        print(f'Статистика за {days_have_passed} {syntax_day(days_have_passed)}')

        nested_list = []
        dc = {}
        create_folder(folder_name_loss)  # создание папки

        print("Сканирование убытка играющих рыцарей:")
        for key1 in pars_dct.keys() & loaded_dict.keys():
            if pars_dct[key1]['gold'] - loaded_dict[key1]['gold'] > 1000 or pars_dct[key1]['victory'] - \
                    loaded_dict[key1]['victory'] > 10:
                print("*", end="")
                url = f'https://s32-ru.battleknight.gameforge.com/common/profile/{key1}/Scores/Player'
                resp = make_request(url, game_sleep=False)
                with open(f'{folder_name_loss}\\{key1}_{pars_dct[key1]["name"]}.html', 'w', encoding='utf-8') as file2:
                    file2.write(resp.text)

                soup = BeautifulSoup(resp.text, 'lxml')
                a = soup.find('table', class_='profileTable').find_all('tr')[4]
                dc[key1] = {"loss": int(a.text.split()[2])}
                with open(STAT_FILE_LOSS, 'rb') as f:
                    loss_dict = pickle.load(f)
                    value_loss = 0 if key1 not in loss_dict else dc[key1]["loss"] - loss_dict[key1]["loss"]
                value1, value2 = pars_dct[key1], loaded_dict[key1]
                profit = value1['gold'] - value2['gold'] if value1['gold'] - value2['gold'] > 0 else 1
                nested_list.append([value1['name'],
                                    value1['clan'],
                                    value1['level'],
                                    value1['gold'] - value2['gold'],
                                    value_loss,
                                    round((100 * value_loss) / profit, 2),
                                    value1['fights'] - value2['fights'],
                                    value1['victory'] - value2['victory'],
                                    value1['defeats'] - value2['defeats']
                                    ])
                sleep(3)
        print()
        print("Сканирование завершено")
        return [nested_list, dc]


def write_2dlist_to_excel(diff_list, write_flag):
    pref = "_all" if write_flag else ""
    excel_file_path = f"bk\\result_xlsx\\stat_{today.day:02d}_{today.month:02d}_{today.year}{pref}.xlsx"

    def set_column_widths(sheet):
        """Устанавливает ширину колонок на основе длины данных."""
        column_widths = [len(str(cell_value)) for row in sheet.iter_rows() for cell_value in row]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[
                get_column_letter(i)].width = column_width + 1  # Добавляем дополнительные пиксели для промежутка

    def add_data_to_sheet(sheet, period, data, header):
        """Добавляет данные и заголовки в лист Excel."""
        sheet.append(period)
        sheet.append(header)
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        for row in data:
            sheet.append(row)
        set_column_widths(sheet)

    # Создаем объект Workbook и листы
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "knight"

    # Добавляем данные на первый лист
    days_have_passed = day(STAT_FILE_NAME)
    data_stat = [f'Статистика за {days_have_passed} {syntax_day(days_have_passed)}:',
                 f'c {DATA_CHANGE_FILE.strftime("%d.%m.%Y %H:%M")}',
                 f'по {datetime.now().strftime("%d.%m.%Y %H:%M")}']
    header1 = ['Имя', 'Орден', 'Уровень', 'Добыча', 'Потери', 'пот/доб (%)', 'Бои', 'Победы', 'Поражения']

    add_data_to_sheet(sheet1, data_stat, diff_list, header1)

    # Обработка данных с использованием pandas
    df = pd.DataFrame(diff_list, columns=header1)
    df['Орден'] = df['Орден'].replace('', 'no orden')
    grouped_df = df.groupby('Орден').agg({'Добыча': 'sum', 'Потери': 'sum'}).reset_index()

    total_dobych = grouped_df['Добыча'].sum()
    grouped_df['доб(%)'] = ((grouped_df['Добыча'] / total_dobych) * 100).round(2)
    grouped_df['пот/доб (%)'] = (grouped_df['Потери'] / grouped_df['Добыча'] * 100).round(2)

    result_list = grouped_df.values.tolist()

    # Создаем второй лист и добавляем данные
    sheet2 = workbook.create_sheet(title="castles")
    header2 = ['Орден', 'Добыча', 'Потери', 'доб(%)', 'пот/доб (%)']
    add_data_to_sheet(sheet2, data_stat, result_list, header2)

    # Сохраняем рабочую книгу
    workbook.save(excel_file_path)

    print(f"Список успешно сохранен в файл {excel_file_path}")
